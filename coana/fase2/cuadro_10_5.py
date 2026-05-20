"""Cuadro 10.5 — Informe de costes primarios por centro de coste.

El *coste primario* de un centro es el total de costes directos e
indirectos (D+I) que se le han asignado en la Fase 1, sin haber sido
imputados desde otro centro o actividad auxiliar. La distinción
directo/indirecto se hace por la columna ``regla_cc`` de las UC:

- ``regla_cc`` nula  → el centro de coste se conoce con exactitud
  desde el dato original (nómina, expediente, apunte con CC explícito,
  etc.). Es un coste *directo*.
- ``regla_cc`` no nula → la asignación al centro la ha hecho un
  criterio de reparto (suministros distribuidos, subcentro/%, par
  centro/%, etc.). Es un coste *indirecto*.

Para cada centro nivel-1 del cuadro 10.4 (8 grupos por finalidad) se
genera una sub-tabla con las filas de la plantilla del cuadro 10.1
(elementos de coste). Las dos filas finales del modelo SUE —
*Centros superiores* (imputaciones desde centros de nivel superior,
Fase 3.a) y *Actividades auxiliares* (Fase 3.d) — quedan a 0 mientras
no se implemente la Fase 3.
"""

from __future__ import annotations

from pathlib import Path

import polars as pl

from coana.fase2._cuadro_jerarquico import persistir_yaml
from coana.fase2.calculo import (
    cargar_árbol_centros_de_coste, cargar_árbol_elementos_de_coste,
    cargar_ucs, descendientes_inclusivo,
)
from coana.fase2.cuadro_10_1 import _PLANTILLA as _PLANTILLA_EC
from coana.fase2.cuadro_10_4 import _PLANTILLA as _PLANTILLA_CC

# Centros sobre los que se emite un sub-cuadro: los 8 nodos nivel-1
# del cuadro 10.4 (clasificación por finalidad).
_CENTROS_PRINCIPALES: list[tuple[str, str, str]] = [
    (cód, slug, nombre)
    for cód, slug, nombre, nivel in _PLANTILLA_CC
    if nivel == 1
]


def generar_cuadro_10_5(ruta_base: Path, dir_informes: Path) -> None:
    ucs = cargar_ucs(ruta_base)
    árbol_ec = cargar_árbol_elementos_de_coste(ruta_base)
    árbol_cc = cargar_árbol_centros_de_coste(ruta_base)

    # Para acelerar: precomputamos un mapeo {ec_slug: set(descendientes)}
    # solo para los slugs de la plantilla de elementos de coste.
    descs_ec: dict[str, set[str]] = {}
    for _, slug, _, _ in _PLANTILLA_EC:
        try:
            descs_ec[slug] = set(descendientes_inclusivo(árbol_ec, slug))
        except KeyError:
            descs_ec[slug] = set()

    # Marca de directo/indirecto: regla_cc nula → directo.
    ucs = ucs.with_columns(
        pl.col("regla_cc").is_null().alias("_es_directo"),
    )

    centros_out: list[dict] = []
    for cód_cc, slug_cc, nombre_cc in _CENTROS_PRINCIPALES:
        try:
            descs_cc = set(descendientes_inclusivo(árbol_cc, slug_cc))
        except KeyError:
            descs_cc = set()
        if not descs_cc:
            continue
        sub = ucs.filter(pl.col("centro_de_coste").is_in(list(descs_cc)))

        filas: list[dict] = []
        total_d = 0.0
        total_i = 0.0
        importe_grupo_d = 0.0
        importe_grupo_i = 0.0
        importe_grupo_p = 0.0
        for cód_ec, slug_ec, nombre_ec, nivel in _PLANTILLA_EC:
            ec_set = descs_ec.get(slug_ec, set())
            sel = sub.filter(pl.col("elemento_de_coste").is_in(list(ec_set)))
            d = float(
                sel.filter(pl.col("_es_directo"))["importe"].sum() or 0.0
            )
            i = float(
                sel.filter(~pl.col("_es_directo"))["importe"].sum() or 0.0
            )
            p = d + i
            if nivel == 1:
                importe_grupo_d = d
                importe_grupo_i = i
                importe_grupo_p = p
                total_d += d
                total_i += i
            filas.append({
                "código": cód_ec,
                "slug": slug_ec,
                "nombre": nombre_ec,
                "nivel": nivel,
                "directo": round(d, 2),
                "indirecto": round(i, 2),
                "primario": round(p, 2),
            })

        total_p = total_d + total_i
        # Filas finales: centros superiores y actividades auxiliares
        # quedan a 0 € hasta que se implemente la Fase 3.
        centros_out.append({
            "código_sue": cód_cc,
            "slug": slug_cc,
            "nombre": nombre_cc,
            "filas": filas,
            "total_coste_primario": {
                "directo": round(total_d, 2),
                "indirecto": round(total_i, 2),
                "primario": round(total_p, 2),
            },
            "centros_superiores": {
                "directo": 0.0, "indirecto": 0.0, "primario": 0.0,
                "pendiente": True,
            },
            "actividades_auxiliares": {
                "directo": 0.0, "indirecto": 0.0, "primario": 0.0,
                "pendiente": True,
            },
            "total": {
                "directo": round(total_d, 2),
                "indirecto": round(total_i, 2),
                "primario": round(total_p, 2),
            },
        })

    datos = {
        "id": "cuadro_10_5",
        "título": (
            "Coste primario de cada centro, con desglose de costes directos e "
            "indirectos de cada elemento de coste"
        ),
        "centros": centros_out,
    }
    persistir_yaml(dir_informes / "cuadro_10_5.yaml", datos)
    _persistir_xlsx(dir_informes / "cuadro_10_5.xlsx", datos)
    n_cs = len(centros_out)
    total = sum(c["total"]["primario"] for c in centros_out)
    print(f"  cuadro_10_5: {n_cs} sub-cuadros, total primario {total:,.2f} €")


def _persistir_xlsx(path: Path, datos: dict) -> None:
    """Hoja única con todos los sub-cuadros encadenados verticalmente."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.styles.borders import Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "cuadro_10_5"
    calibri = Font(name="Calibri", size=11)
    calibri_bold = Font(name="Calibri", size=11, bold=True)
    calibri_title = Font(name="Calibri", size=13, bold=True)
    calibri_sub = Font(name="Calibri", size=12, bold=True)
    fill_grupo = PatternFill("solid", fgColor="E7EEF6")
    fill_sub = PatternFill("solid", fgColor="D6E2F2")
    fill_pendiente = PatternFill("solid", fgColor="F2F2F2")
    thin = Side(style="thin", color="888888")
    medio = Side(style="medium", color="333333")
    border_filas = Border(bottom=thin)

    ws["A1"] = datos["título"]
    ws["A1"].font = calibri_title
    ws.merge_cells("A1:D1")

    row = 3
    for centro in datos["centros"]:
        # Subtítulo del sub-cuadro.
        c = ws.cell(row=row, column=1, value=f"{centro['código_sue']} — {centro['nombre']}")
        c.font = calibri_sub
        c.fill = fill_sub
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        row += 1

        headers = ["Código", "Elemento de coste", "Directo (€)", "Indirecto (€)", "Primario (D+I) (€)"]
        # 5 columnas; ajustamos
        for col, h in enumerate(headers, start=1):
            cc = ws.cell(row=row, column=col, value=h)
            cc.font = calibri_bold
            cc.border = Border(top=medio, bottom=medio)
            cc.alignment = Alignment(horizontal="left" if col <= 2 else "right")
        row += 1

        for fila in centro["filas"]:
            es_grupo = fila["nivel"] == 1
            sangría = "" if es_grupo else "    " * (fila["nivel"] - 1)
            font = calibri_bold if es_grupo else calibri
            ws.cell(row=row, column=1, value=fila["código"]).font = font
            ws.cell(row=row, column=2, value=f"{sangría}{fila['nombre']}").font = font
            for off, key in enumerate(("directo", "indirecto", "primario"), start=3):
                cc = ws.cell(row=row, column=off, value=fila[key])
                cc.number_format = '#,##0.00'
                cc.font = font
            if es_grupo:
                for col in range(1, 6):
                    ws.cell(row=row, column=col).fill = fill_grupo
            for col in range(1, 6):
                ws.cell(row=row, column=col).border = border_filas
            row += 1

        # Filas finales.
        for etiqueta, valores, pendiente in (
            ("Total coste primario", centro["total_coste_primario"], False),
            ("Centros superiores (Fase 3.a, pendiente)", centro["centros_superiores"], True),
            ("Actividades auxiliares (Fase 3.d, pendiente)", centro["actividades_auxiliares"], True),
            ("Total", centro["total"], False),
        ):
            ws.cell(row=row, column=1, value="").font = calibri_bold
            ws.cell(row=row, column=2, value=etiqueta).font = calibri_bold
            for off, key in enumerate(("directo", "indirecto", "primario"), start=3):
                cc = ws.cell(row=row, column=off, value=valores[key])
                cc.number_format = '#,##0.00'
                cc.font = calibri_bold
            if pendiente:
                for col in range(1, 6):
                    ws.cell(row=row, column=col).fill = fill_pendiente
            for col in range(1, 6):
                ws.cell(row=row, column=col).border = Border(top=thin)
            row += 1
        # Separación entre centros.
        row += 1

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 75
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 22

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
