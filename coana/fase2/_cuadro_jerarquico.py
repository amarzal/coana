"""Motor común para cuadros tipo plantilla SUE.

Un cuadro jerárquico se compone de:
- Una *plantilla* fija con la lista de filas a mostrar:
  ``(código_SUE, slug, nombre, nivel)``.
- Un *árbol* sobre el que se computa ``A(X)+B(X)`` (suma del subárbol)
  para cada slug del cuadro.
- Una *columna* de las UC que enlaza con el árbol
  (``elemento_de_coste``, ``centro_de_coste`` o ``actividad``).

Calcula ``importe``, ``% elemento`` (sobre el grupo nivel-1 más
próximo) y ``% total`` (solo en filas nivel-1). Si hay UC cuyo
slug no cae bajo ninguna rama de la plantilla, añade una fila final
«Sin clasificar en plantilla SUE» antes del total.

Persiste un YAML y un XLSX (Calibri).
"""

from __future__ import annotations

from pathlib import Path

import polars as pl
import yaml

from coana.fase2.calculo import importes_por_nodo, total_general
from coana.util import Árbol


def generar_cuadro_jerárquico(
    *,
    id_cuadro: str,
    título: str,
    encabezado_concepto: str,
    plantilla: list[tuple[str, str, str, int]],
    ucs: pl.DataFrame,
    árbol: Árbol,
    columna_coordenada: str,
    dir_informes: Path,
    expandir_nivel_3: set[str] | None = None,
) -> dict:
    """Computa y persiste el cuadro. Devuelve los datos calculados.

    Si se pasa ``expandir_nivel_3``, después de cada fila nivel-2 cuyo
    slug esté en el conjunto se insertan filas nivel-3 con los hijos
    directos del nodo en el árbol (códigos derivados del padre, p. ej.
    ``01.01`` → ``01.01.01``, ``01.01.02``…). El ``% elemento`` de las
    filas nivel-3 sigue siendo relativo al grupo nivel-1 (coherente
    con el cuadro 10.4 oficial: la suma de los nietos bajo un padre
    nivel-2 reproduce el ``% elemento`` del propio nivel-2).
    """
    expandir = expandir_nivel_3 or set()
    importes = importes_por_nodo(ucs, árbol, columna_coordenada)
    total = total_general(ucs)

    filas: list[dict] = []
    importe_grupo_actual: float = 0.0
    for cód, slug, nombre, nivel in plantilla:
        importe = importes.get(slug, 0.0)
        if nivel == 1:
            importe_grupo_actual = importe
            pct_elemento = 100.0 if importe != 0 else 0.0
            pct_total = (importe / total * 100.0) if total else 0.0
        else:
            pct_elemento = (
                (importe / importe_grupo_actual * 100.0)
                if importe_grupo_actual else 0.0
            )
            pct_total = None
        filas.append({
            "código": cód,
            "slug": slug,
            "nombre": nombre,
            "nivel": nivel,
            "importe": round(importe, 2),
            "pct_elemento": round(pct_elemento, 4),
            "pct_total": (round(pct_total, 4) if pct_total is not None else None),
        })
        # Expansión a nivel-3: hijos directos del nodo del árbol.
        if nivel == 2 and slug in expandir:
            try:
                nodo = árbol._nodo(slug)
            except KeyError:
                continue
            for i, hijo in enumerate(nodo.hijos, start=1):
                slug_h = hijo.identificador
                if not slug_h:
                    continue
                importe_h = importes.get(slug_h, 0.0)
                pct_e_h = (
                    (importe_h / importe_grupo_actual * 100.0)
                    if importe_grupo_actual else 0.0
                )
                filas.append({
                    "código": f"{cód}.{i:02d}",
                    "slug": slug_h,
                    "nombre": hijo.descripción,
                    "nivel": 3,
                    "importe": round(importe_h, 2),
                    "pct_elemento": round(pct_e_h, 4),
                    "pct_total": None,
                })

    # Fila «sin clasificar» si hay UC cuyo slug no cae bajo ninguno de
    # los nodos nivel-1 listados.
    slugs_nivel_1 = {s for c, s, _, n in plantilla if n == 1}
    cubierto = sum(importes.get(s, 0.0) for s in slugs_nivel_1)
    no_clasificable = round(total - cubierto, 2)
    if abs(no_clasificable) >= 0.01:
        filas.append({
            "código": "—",
            "slug": "_no_clasificable",
            "nombre": "Sin clasificar en plantilla SUE",
            "nivel": 1,
            "importe": no_clasificable,
            "pct_elemento": 100.0,
            "pct_total": round(no_clasificable / total * 100.0, 4) if total else 0.0,
        })

    datos = {
        "id": id_cuadro,
        "título": título,
        "encabezado_concepto": encabezado_concepto,
        "total": round(total, 2),
        "filas": filas,
    }
    persistir_yaml(dir_informes / f"{id_cuadro}.yaml", datos)
    persistir_xlsx_jerárquico(dir_informes / f"{id_cuadro}.xlsx", datos)
    extra = (
        f" (no clasificable {no_clasificable:,.2f} €)"
        if abs(no_clasificable) >= 0.01 else ""
    )
    print(f"  {id_cuadro}: {len(filas):,} filas, total {total:,.2f} €{extra}")
    return datos


def persistir_yaml(path: Path, datos: dict) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as f:
        yaml.safe_dump(datos, f, allow_unicode=True, sort_keys=False)


def persistir_xlsx_jerárquico(path: Path, datos: dict) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.styles.borders import Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = datos["id"]
    calibri = Font(name="Calibri", size=11)
    calibri_bold = Font(name="Calibri", size=11, bold=True)
    calibri_title = Font(name="Calibri", size=13, bold=True)
    fill_grupo = PatternFill("solid", fgColor="E7EEF6")
    thin = Side(style="thin", color="888888")
    medio = Side(style="medium", color="333333")
    border_filas = Border(bottom=thin)

    ws["A1"] = datos["título"]
    ws["A1"].font = calibri_title
    ws.merge_cells("A1:E1")

    headers = ["Código", datos["encabezado_concepto"], "Importe", "% elemento", "% total"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=3, column=col, value=h)
        c.font = calibri_bold
        c.border = Border(top=medio, bottom=medio)
        c.alignment = Alignment(horizontal="left" if col <= 2 else "right")

    row = 4
    for fila in datos["filas"]:
        cód = fila["código"]
        nombre = fila["nombre"]
        importe = fila["importe"]
        pct_e = fila["pct_elemento"]
        pct_t = fila["pct_total"]
        es_grupo = fila["nivel"] == 1
        sangría = "" if es_grupo else "    " * (fila["nivel"] - 1)
        font = calibri_bold if es_grupo else calibri
        ws.cell(row=row, column=1, value=cód).font = font
        ws.cell(row=row, column=2, value=f"{sangría}{nombre}").font = font
        ce = ws.cell(row=row, column=3, value=importe)
        ce.number_format = '#,##0.00\\ "€"'
        ce.font = font
        cpe = ws.cell(
            row=row, column=4,
            value=(pct_e / 100.0) if pct_e is not None else None,
        )
        cpe.number_format = "0.00%"
        cpe.font = font
        cpt = ws.cell(
            row=row, column=5,
            value=(pct_t / 100.0) if pct_t is not None else None,
        )
        cpt.number_format = "0.00%"
        cpt.font = font
        if es_grupo:
            for col in range(1, 6):
                ws.cell(row=row, column=col).fill = fill_grupo
        for col in range(1, 6):
            ws.cell(row=row, column=col).border = border_filas
        row += 1

    # Fila Total.
    ws.cell(row=row, column=1, value="").font = calibri_bold
    ws.cell(row=row, column=2, value="Total").font = calibri_bold
    ct = ws.cell(row=row, column=3, value=datos["total"])
    ct.number_format = '#,##0.00\\ "€"'
    ct.font = calibri_bold
    ws.cell(row=row, column=4, value=None)
    cpt = ws.cell(row=row, column=5, value=1.0)
    cpt.number_format = "0.00%"
    cpt.font = calibri_bold
    for col in range(1, 6):
        ws.cell(row=row, column=col).border = Border(top=medio, bottom=medio)
        ws.cell(row=row, column=col).font = calibri_bold

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 75
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 13
    ws.column_dimensions["E"].width = 13

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
