"""Cuadro 10.7 — Composición del coste de las actividades finalistas.

Matriz «actividad finalista × tipo de centro». Filas: jerarquía de
actividades finalistas según la plantilla del PDF (códigos 01, 01.01,
01.01.01, …). Columnas: agrupaciones de centros de coste:

- Depts.        → subárbol `departamentos`.
- Biblioteca    → subárbol `bibliotecas`.
- Laboratorios  → `laboratorios-investigación` + `laboratorios-docencia-investigación`.
- Aulas         → `aulas-laboratorios-docentes`.
- DAG           → centros directivos / administrativos / generales:
                  `soporte`, `apoyo-docencia-investigación` y `anexos`.
- Otros         → cualquier otro centro de coste (institutos,
                  cátedras, otros centros docentes/investigación, etc.).
- Total         → suma horizontal.

El cuadro solo cuenta las UC cuya actividad cuelga del subárbol
`principales` del árbol de actividades. Las actividades DAG, anexas,
organización, TRUPI, subactividad y almacenaje no aparecen.
"""

from __future__ import annotations

from pathlib import Path

import polars as pl

from coana.fase2._cuadro_jerarquico import persistir_yaml
from coana.fase2.calculo import (
    cargar_árbol_actividades, cargar_árbol_centros_de_coste, cargar_ucs,
    descendientes_inclusivo,
)

# (código_SUE, slug_árbol, nombre_oficial, nivel)
_PLANTILLA: list[tuple[str, str, str, int]] = [
    ("01",             "principales",                "Actividades finalistas o principales y asimiladas", 1),
    ("01.01",          "docencia",                   "Actividades de docencia", 2),
    ("01.01.01",       "estudios-oficiales",         "Estudios oficiales", 3),
    ("01.01.02",       "enseñanzas-propias",         "Enseñanzas propias universitarias", 3),
    ("01.02",          "ai",                         "Actividades de investigación y transferencia del conocimiento", 2),
    ("01.02.01",       "ai-financiación-propia",     "Actividades de investigación con financiación propia", 3),
    ("01.02.02",       "ait-financiación-externa",   "Actividades de investigación y transferencia con financiación externa", 3),
    ("01.02.02.01",    "ai-financiación-competitiva", "Actividades financiadas por agentes externos en convocatorias competitivas", 4),
    ("01.02.02.01.01", "ai-regional",                "Actividades con financiación de programas regionales", 5),
    ("01.02.02.01.02", "ai-nacional",                "Actividades con financiación de programas nacionales", 5),
    ("01.02.02.01.03", "ai-internacional",           "Actividades con financiación de programas internacionales", 5),
    ("01.02.02.01.04", "ai-otras-competitivas",      "Otras actividades de investigación en convocatorias competitivas", 5),
    ("01.02.02.02",    "transf",                     "Actividades de investigación aplicada y transferencia", 4),
    ("01.02.03",       "doctorado",                  "Doctorado", 3),
    ("01.03",          "deportes-extensión-universitaria", "Actividades deportivas y de extensión universitaria", 2),
]

# (id_columna, nombre, [slugs_raíz_del_grupo])
_COLUMNAS: list[tuple[str, str, list[str]]] = [
    ("depts",        "Depts.",       ["departamentos"]),
    ("biblioteca",   "Biblioteca",   ["bibliotecas"]),
    ("laboratorios", "Laboratorios", ["laboratorios-investigación", "laboratorios-docencia-investigación"]),
    ("aulas",        "Aulas",        ["aulas-laboratorios-docentes"]),
    ("dag",          "DAG",          ["soporte", "apoyo-docencia-investigación", "anexos"]),
]


def generar_cuadro_10_7(ruta_base: Path, dir_informes: Path) -> None:
    ucs = cargar_ucs(ruta_base)
    árbol_act = cargar_árbol_actividades(ruta_base)
    árbol_cc = cargar_árbol_centros_de_coste(ruta_base)

    # Pre-cómputo: descendientes inclusivos de cada actividad de la plantilla.
    descs_act: dict[str, set[str]] = {}
    for _, slug, _, _ in _PLANTILLA:
        try:
            descs_act[slug] = set(descendientes_inclusivo(árbol_act, slug))
        except KeyError:
            descs_act[slug] = set()

    # Pre-cómputo: descendientes inclusivos de cada grupo de columnas.
    descs_col: dict[str, set[str]] = {}
    cubiertos_cc: set[str] = set()
    for id_col, _, raíces in _COLUMNAS:
        s: set[str] = set()
        for raíz in raíces:
            try:
                s |= set(descendientes_inclusivo(árbol_cc, raíz))
            except KeyError:
                pass
        descs_col[id_col] = s
        cubiertos_cc |= s

    filas_out: list[dict] = []
    for cód, slug_act, nombre, nivel in _PLANTILLA:
        act_set = descs_act.get(slug_act, set())
        if not act_set:
            valores = {id_col: 0.0 for id_col, _, _ in _COLUMNAS}
            valores["otros"] = 0.0
            valores["total"] = 0.0
            filas_out.append({
                "código": cód, "slug": slug_act, "nombre": nombre,
                "nivel": nivel, "valores": valores,
            })
            continue
        sub_ucs = ucs.filter(pl.col("actividad").is_in(list(act_set)))
        valores: dict[str, float] = {}
        importe_total_fila = 0.0
        for id_col, _, _ in _COLUMNAS:
            cc_set = descs_col[id_col]
            sel = sub_ucs.filter(pl.col("centro_de_coste").is_in(list(cc_set)))
            v = float(sel["importe"].sum() or 0.0)
            valores[id_col] = round(v, 2)
            importe_total_fila += v
        # «Otros»: lo que cae en la actividad pero su CC no está en
        # ningún grupo de columna definido.
        sel_otros = sub_ucs.filter(~pl.col("centro_de_coste").is_in(list(cubiertos_cc)))
        otros = float(sel_otros["importe"].sum() or 0.0)
        valores["otros"] = round(otros, 2)
        importe_total_fila += otros
        valores["total"] = round(importe_total_fila, 2)
        filas_out.append({
            "código": cód, "slug": slug_act, "nombre": nombre,
            "nivel": nivel, "valores": valores,
        })

    # Total de la columna (suma de filas nivel-1, que son las únicas
    # que cubren la masa entera sin doble cómputo).
    nivel_1_filas = [f for f in filas_out if f["nivel"] == 1]
    columnas_total: dict[str, float] = {}
    for id_col in [c[0] for c in _COLUMNAS] + ["otros", "total"]:
        columnas_total[id_col] = round(
            sum(f["valores"][id_col] for f in nivel_1_filas), 2
        )

    columnas_meta = [
        {"id": id_col, "nombre": nombre}
        for id_col, nombre, _ in _COLUMNAS
    ] + [
        {"id": "otros", "nombre": "Otros"},
        {"id": "total", "nombre": "Total"},
    ]

    datos = {
        "id": "cuadro_10_7",
        "título": "Composición del coste de las actividades finalistas",
        "columnas": columnas_meta,
        "filas": filas_out,
        "total": columnas_total,
    }
    persistir_yaml(dir_informes / "cuadro_10_7.yaml", datos)
    _persistir_xlsx(dir_informes / "cuadro_10_7.xlsx", datos)
    print(
        f"  cuadro_10_7: {len(filas_out)} filas × {len(columnas_meta)} cols, "
        f"total finalistas {columnas_total['total']:,.2f} €"
    )


def _persistir_xlsx(path, datos):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.styles.borders import Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "cuadro_10_7"
    calibri = Font(name="Calibri", size=11)
    calibri_bold = Font(name="Calibri", size=11, bold=True)
    calibri_title = Font(name="Calibri", size=13, bold=True)
    fill_grupo = PatternFill("solid", fgColor="E7EEF6")
    thin = Side(style="thin", color="888888")
    medio = Side(style="medium", color="333333")

    columnas = datos["columnas"]
    ncols_data = len(columnas)

    ws["A1"] = datos["título"]
    ws["A1"].font = calibri_title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=1 + ncols_data)

    headers = ["Actividad"] + [f"{c['nombre']} (€)" for c in columnas]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=3, column=col, value=h)
        c.font = calibri_bold
        c.border = Border(top=medio, bottom=medio)
        c.alignment = Alignment(horizontal="left" if col == 1 else "right")

    row = 4
    for fila in datos["filas"]:
        es_grupo = fila["nivel"] == 1
        sangría = "" if es_grupo else "    " * (fila["nivel"] - 1)
        font = calibri_bold if es_grupo else calibri
        ws.cell(row=row, column=1, value=f"{sangría}{fila['código']} {fila['nombre']}").font = font
        for off, c_meta in enumerate(columnas, start=2):
            v = fila["valores"].get(c_meta["id"], 0.0)
            cc = ws.cell(row=row, column=off, value=v)
            cc.number_format = '#,##0.00'
            cc.font = font
        if es_grupo:
            for col in range(1, 2 + ncols_data):
                ws.cell(row=row, column=col).fill = fill_grupo
        for col in range(1, 2 + ncols_data):
            ws.cell(row=row, column=col).border = Border(bottom=thin)
        row += 1

    # Fila Total.
    ws.cell(row=row, column=1, value="Total").font = calibri_bold
    for off, c_meta in enumerate(columnas, start=2):
        v = datos["total"].get(c_meta["id"], 0.0)
        cc = ws.cell(row=row, column=off, value=v)
        cc.number_format = '#,##0.00'
        cc.font = calibri_bold
    for col in range(1, 2 + ncols_data):
        ws.cell(row=row, column=col).border = Border(top=medio, bottom=medio)

    ws.column_dimensions["A"].width = 70
    for off in range(ncols_data):
        ws.column_dimensions[chr(ord("B") + off)].width = 16

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
