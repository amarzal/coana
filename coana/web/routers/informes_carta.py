"""Informes a la carta: agrupación jerárquica de UC por CC/ACT/EC.

Endpoints adicionales: exportación a XLSX y PDF del resultado.

El usuario selecciona uno o más slugs de cada uno de los tres árboles
(centros de coste, actividades, elementos de coste), elige el orden
de los tres niveles (p. ej. CC → ACT → EC) y obtiene un listado
jerárquico con conteo de UC e importe acumulado por nodo.

Selección por subárbol: cada slug seleccionado incluye implícitamente
todos sus descendientes en el árbol correspondiente. Lista vacía =
«todos los slugs».

Configuraciones guardadas: `data/informes/carta_configs/{nombre}.yaml`.
"""

from __future__ import annotations

import io
import re
import subprocess
import tempfile
from pathlib import Path
from typing import Any

import polars as pl
import yaml
from fastapi import APIRouter, HTTPException
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field

from coana.fase2.calculo import (
    cargar_árbol_actividades, cargar_árbol_centros_de_coste,
    cargar_árbol_elementos_de_coste, cargar_ucs, descendientes_inclusivo,
)
from coana.util import Árbol

router = APIRouter()


_RUTA_BASE = Path("data")
_DIR_CONFIGS = _RUTA_BASE / "informes" / "carta_configs"
_PATH_POST_REPARTO = _RUTA_BASE / "fase1" / "reparto" / "uc_post_reparto.parquet"


def _cargar_ucs() -> pl.DataFrame:
    """Conjunto de UC del informe: el de **después del reparto de
    actividades dag** (`uc_post_reparto.parquet`), donde los costes
    indirectos (dag) ya están repercutidos a las actividades finalistas,
    de modo que una actividad finalista (p. ej. `grado-cafe`) incluye su
    parte dag. Si el reparto aún no se ha ejecutado, se cae al conjunto
    directo de la Fase 1 (`unidades de coste.xlsx`)."""
    if _PATH_POST_REPARTO.exists():
        return pl.read_parquet(_PATH_POST_REPARTO)
    return cargar_ucs(_RUTA_BASE)


class Filtro(BaseModel):
    # Eje que vertebra el informe (árbol monográfico): "cc" | "act" | "ec".
    estructura: str = "act"
    centros_de_coste: list[str] = Field(default_factory=list)
    actividades: list[str] = Field(default_factory=list)
    elementos_de_coste: list[str] = Field(default_factory=list)
    # Campos heredados (ya no se usan: el informe es monográfico). Se conservan
    # opcionales para no romper configuraciones YAML guardadas antes del cambio.
    orden: list[str] = Field(default_factory=lambda: ["cc", "act", "ec"])
    agregado: dict[str, bool] = Field(
        default_factory=lambda: {"cc": True, "act": True, "ec": True}
    )


class Opciones(BaseModel):
    centros_de_coste: list[dict]
    actividades: list[dict]
    elementos_de_coste: list[dict]


def _opciones_de(árbol: Árbol) -> list[dict]:
    """Lista plana de nodos del árbol, ordenados por código."""
    out: list[dict] = []
    for ident, nodo in árbol._por_id.items():
        if not ident or nodo is árbol.raíz:
            continue
        out.append({
            "slug": ident,
            "código": nodo.código,
            "descripción": nodo.descripción,
        })
    out.sort(key=lambda x: x["código"])
    return out


@router.get("/opciones", response_model=Opciones)
def opciones() -> Opciones:
    """Devuelve los slugs disponibles en cada árbol para los selectores."""
    return Opciones(
        centros_de_coste=_opciones_de(cargar_árbol_centros_de_coste(_RUTA_BASE)),
        actividades=_opciones_de(cargar_árbol_actividades(_RUTA_BASE)),
        elementos_de_coste=_opciones_de(cargar_árbol_elementos_de_coste(_RUTA_BASE)),
    )


def _expandir(slugs: list[str], árbol: Árbol) -> set[str]:
    """Conjunto de slugs efectivos: para cada `s` en `slugs`, los
    descendientes inclusivos en `árbol`. Lista vacía = no filtrar."""
    if not slugs:
        return set()
    s: set[str] = set()
    for slug in slugs:
        try:
            s |= set(descendientes_inclusivo(árbol, slug))
        except KeyError:
            pass
    return s


class NodoJerarquía(BaseModel):
    nivel: int
    eje: str       # "cc" | "act" | "ec"
    slug: str
    código: str
    descripción: str
    # `n_ucs`/`importe` son el TOTAL del subárbol (directo + descendientes). El
    # «directo» (a) es lo asignado exactamente a este slug; los «descendientes»
    # (b) = total − directo. Los «ancestros» (c) = `importe_ancestros`: la suma
    # de las fracciones de las UC de los ancestros (camino a la raíz) que le
    # corresponden a este nodo por roll-down step-down (proporcional al coste
    # directo del subárbol). El total cargado de la fila = importe + importe_ancestros.
    n_ucs: int
    importe: float
    n_ucs_directo: int = 0
    importe_directo: float = 0.0
    importe_ancestros: float = 0.0
    n_ucs_ancestros: int = 0
    hijos: list["NodoJerarquía"] = Field(default_factory=list)


NodoJerarquía.model_rebuild()


class Resultado(BaseModel):
    orden: list[str]
    n_ucs: int
    importe: float
    raíces: list[NodoJerarquía]


_EJES_COLS = {
    "cc": "centro_de_coste",
    "act": "actividad",
    "ec": "elemento_de_coste",
}
_EJES_ÁRBOL_FN = {
    "cc": cargar_árbol_centros_de_coste,
    "act": cargar_árbol_actividades,
    "ec": cargar_árbol_elementos_de_coste,
}


_RAÍZ_LABEL = {
    "cc": "Todos los centros de coste",
    "act": "Todas las actividades",
    "ec": "Todos los elementos de coste",
}


def _meta_nodo(árbol: Árbol, eje: str, slug: str) -> tuple[str, str]:
    """`(código, descripción)` de un slug. La raíz virtual recibe una etiqueta
    legible; un slug ausente del árbol (huérfano) recibe código `?`."""
    try:
        raíz_id = árbol.raíz.identificador
    except AttributeError:
        raíz_id = None
    if raíz_id is not None and slug == raíz_id:
        return ("—", _RAÍZ_LABEL.get(eje, "Todos"))
    try:
        nodo = árbol._nodo(slug)
        return (nodo.código, nodo.descripción)
    except KeyError:
        return ("?", slug)


def _construir_estructura(
    df: pl.DataFrame, eje: str, árbol: Árbol, foco_slugs: list[str],
) -> list[NodoJerarquía]:
    """Árbol *monográfico* de un eje, anclado en la raíz real (UJI).

    `foco_slugs` es solo *foco*: sin foco se muestra el árbol entero; con foco se
    muestran la *espina* de ancestros desde la raíz hasta cada nodo + su subárbol
    completo. Los importes son el coste *total real* del nodo: directo (a),
    descendientes de todo su subárbol (b) y la fracción que le baja de sus
    ancestros por roll-down step-down proporcional al coste directo del subárbol
    (c) — calculado sobre el árbol completo. No se mezclan estructuras: el árbol
    es de un único eje."""
    col = _EJES_COLS[eje]
    raíz_id = árbol.raíz.identificador
    raíces_ámbito = {raíz_id}
    agg = df.group_by(col).agg(
        pl.len().alias("_n"), pl.col("importe").sum().alias("_imp"),
    )
    directo_n: dict[str, int] = {}
    directo_imp: dict[str, float] = {}
    for r in agg.iter_rows(named=True):
        s = r[col]
        if s is None:
            continue
        directo_n[s] = int(r["_n"])
        directo_imp[s] = float(r["_imp"] or 0.0)

    # Slugs que forman la estructura del árbol: los que tienen coste directo y,
    # si hay foco, también la espina + subárbol de cada nodo enfocado tomados
    # del árbol real. Así un nodo enfocado SIN coste (p. ej. una actividad dag,
    # cuyo coste ya se repartió a las finalistas) sigue apareciendo en su sitio
    # en vez de perderse y dejar que la vista colapse a la raíz con el total.
    slugs_estructura: set[str] = set(directo_n)
    for s in foco_slugs:
        slugs_estructura.update(_camino_árbol(árbol, s, raíces_ámbito))
        slugs_estructura.update(descendientes_inclusivo(árbol, s))

    hijos_de: dict[str, set[str]] = {}
    raíz_nodos: set[str] = set()
    for s in slugs_estructura:
        cam = _camino_árbol(árbol, s, raíces_ámbito)
        raíz_nodos.add(cam[0])
        for padre, hijo in zip(cam, cam[1:]):
            hijos_de.setdefault(padre, set()).add(hijo)

    # Coste/nº de subárbol (b) y roll-down de ancestros (c).
    subtree_d: dict[str, float] = {}
    subtree_n: dict[str, int] = {}

    def _subtree(slug: str) -> None:
        s = directo_imp.get(slug, 0.0)
        n = directo_n.get(slug, 0)
        for c in hijos_de.get(slug, set()):
            _subtree(c)
            s += subtree_d[c]
            n += subtree_n[c]
        subtree_d[slug] = s
        subtree_n[slug] = n

    for r in raíz_nodos:
        _subtree(r)

    recibido: dict[str, float] = {}
    recibido_n: dict[str, int] = {}

    def _bajar(slug: str, rec: float, anc_n: int) -> None:
        recibido[slug] = rec
        recibido_n[slug] = anc_n
        salida = directo_imp.get(slug, 0.0) + rec
        hijos = hijos_de.get(slug, set())
        if hijos:
            total = sum(subtree_d[c] for c in hijos)
            anc_n_hijos = anc_n + directo_n.get(slug, 0)
            for c in hijos:
                # Si el subárbol no tiene coste (todos los hijos a 0, p. ej. una
                # rama dag ya repartida), no hay nada que bajar.
                frac = salida * subtree_d[c] / total if total > 0 else 0.0
                _bajar(c, frac, anc_n_hijos)

    for r in raíz_nodos:
        _bajar(r, 0.0, 0)

    # Nodos a mostrar (foco): sin foco, todo; con foco, espina + subárbol.
    mostrar: set[str] | None
    if not foco_slugs:
        mostrar = None
    else:
        mostrar = set()
        for s in foco_slugs:
            for x in _camino_árbol(árbol, s, raíces_ámbito):
                mostrar.add(x)
            pila = [s]
            while pila:
                x = pila.pop()
                mostrar.add(x)
                pila.extend(hijos_de.get(x, set()))

    def _hijos_mostrados(slug: str) -> set[str]:
        h = hijos_de.get(slug, set())
        return h if mostrar is None else (h & mostrar)

    def _build(slug: str) -> NodoJerarquía:
        cod, desc = _meta_nodo(árbol, eje, slug)
        hijos_tree = [
            _build(c)
            for c in sorted(_hijos_mostrados(slug), key=lambda x: _meta_nodo(árbol, eje, x)[0])
        ]
        dn = directo_n.get(slug, 0)
        di = directo_imp.get(slug, 0.0)
        return NodoJerarquía(
            nivel=1, eje=eje, slug=slug, código=cod, descripción=desc,
            n_ucs=subtree_n.get(slug, dn),
            importe=round(subtree_d.get(slug, di), 2),
            n_ucs_directo=dn, importe_directo=round(di, 2),
            importe_ancestros=round(recibido.get(slug, 0.0), 2),
            n_ucs_ancestros=recibido_n.get(slug, 0),
            hijos=hijos_tree,
        )

    return [
        _build(r) for r in sorted(raíz_nodos, key=lambda x: _meta_nodo(árbol, eje, x)[0])
    ]


@router.post("/consulta", response_model=Resultado)
def consulta(filtro: Filtro) -> Resultado:
    """Informe *monográfico*: una sola estructura (eje `filtro.estructura`) como
    árbol, anclado en la raíz real. Los otros dos ejes actúan SOLO como filtro
    (subárbol de su selección); la selección del eje de estructura es solo foco."""
    estructura = filtro.estructura
    if estructura not in _EJES_COLS:
        raise HTTPException(
            status_code=400,
            detail=f"`estructura` debe ser 'cc', 'act' o 'ec': {estructura!r}",
        )
    ucs = _cargar_ucs()
    árbol = _EJES_ÁRBOL_FN[estructura](_RUTA_BASE)
    sel_eje = {
        "cc": filtro.centros_de_coste,
        "act": filtro.actividades,
        "ec": filtro.elementos_de_coste,
    }

    # Filtrar por los OTROS ejes (subárbol de su selección). El eje de estructura
    # no filtra: su selección es solo foco de la vista.
    sub = ucs
    for eje in ("cc", "act", "ec"):
        if eje == estructura:
            continue
        sel = _expandir(sel_eje[eje], _EJES_ÁRBOL_FN[eje](_RUTA_BASE))
        if sel:
            sub = sub.filter(pl.col(_EJES_COLS[eje]).is_in(list(sel)))

    if sub.is_empty():
        return Resultado(orden=[estructura], n_ucs=0, importe=0.0, raíces=[])

    total_n = int(sub.height)
    total_imp = float(sub["importe"].sum() or 0.0)
    raíces = _construir_estructura(sub, estructura, árbol, sel_eje[estructura])
    return Resultado(
        orden=[estructura],
        n_ucs=total_n,
        importe=round(total_imp, 2),
        raíces=raíces,
    )


class PeticiónUcs(BaseModel):
    centros_de_coste: list[str] = Field(default_factory=list)
    actividades: list[str] = Field(default_factory=list)
    elementos_de_coste: list[str] = Field(default_factory=list)
    # Para ese eje ("cc"|"act"|"ec"): `exacto_eje` filtra por slug EXACTO (UC
    # asignadas directamente al nodo); `indirecto_eje` filtra por el subárbol
    # EXCLUYENDO el slug exacto (UC que el nodo recibe de sus descendientes).
    exacto_eje: str | None = None
    indirecto_eje: str | None = None
    limit: int = 5000


@router.post("/uc")
def uc_de_combinación(p: PeticiónUcs) -> dict[str, Any]:
    """Lista de UC que cuelgan de los slugs indicados (expansión por subárbol;
    `exacto_eje` = solo el slug exacto; `indirecto_eje` = subárbol sin el slug
    exacto). Hasta `limit` filas."""
    ucs = _cargar_ucs()

    def _slugs(eje: str, slugs: list[str], árbol: Árbol) -> set[str]:
        if p.exacto_eje == eje:
            return set(slugs)
        if p.indirecto_eje == eje:
            return _expandir(slugs, árbol) - set(slugs)
        return _expandir(slugs, árbol)

    sels = {
        "cc": _slugs("cc", p.centros_de_coste, cargar_árbol_centros_de_coste(_RUTA_BASE)),
        "act": _slugs("act", p.actividades, cargar_árbol_actividades(_RUTA_BASE)),
        "ec": _slugs("ec", p.elementos_de_coste, cargar_árbol_elementos_de_coste(_RUTA_BASE)),
    }
    # Los ejes con modo exacto/indirecto se filtran SIEMPRE (aunque el conjunto
    # sea vacío → 0 filas), p. ej. el indirecto de un nodo hoja. El resto, solo
    # si hay selección (conjunto vacío = no filtrar = todos).
    forzados = {p.exacto_eje, p.indirecto_eje} - {None}
    sub = ucs
    for eje, col in _EJES_COLS.items():
        if sels[eje] or eje in forzados:
            sub = sub.filter(pl.col(col).is_in(list(sels[eje])))
    n_total = int(sub.height)
    if "importe" in sub.columns:
        sub = sub.sort(pl.col("importe").abs(), descending=True)
    filas = sub.head(p.limit).to_dicts()
    return {
        "n_total": n_total,
        "n_devueltas": len(filas),
        "filas": filas,
    }


def _camino_árbol(árbol: Árbol, slug: str, raíces: set[str]) -> list[str]:
    """Identificadores desde la raíz de ámbito (incluida) hasta `slug` (incluido)."""
    try:
        nodo = árbol._nodo(slug)
    except KeyError:
        return [slug]
    cadena: list[str] = []
    while nodo is not None:
        cadena.append(nodo.identificador)
        if nodo.identificador in raíces:
            break
        nodo = nodo.padre
    cadena.reverse()
    return cadena


class PeticiónAncestros(BaseModel):
    # Contexto de los OTROS ejes (subárbol del slug de contexto). El eje del
    # nodo se ignora aquí: su ámbito lo fija `scope_slugs`.
    centros_de_coste: list[str] = Field(default_factory=list)
    actividades: list[str] = Field(default_factory=list)
    elementos_de_coste: list[str] = Field(default_factory=list)
    eje: str            # "cc" | "act" | "ec": eje del nodo N
    slug: str           # nodo N
    # Ámbito del eje del nodo (selección del informe en ese eje); vacío = todo el árbol.
    scope_slugs: list[str] = Field(default_factory=list)
    limit: int = 5000


@router.post("/uc-ancestros")
def uc_ancestros(p: PeticiónAncestros) -> dict[str, Any]:
    """UC de los ancestros estrictos de `slug` (en su eje), cada una con la
    *fracción* (`_fraccion`) que le corresponde a `slug` por roll-down step-down
    (proporcional al coste directo del subárbol). El importe efectivamente
    aportado al nodo es `importe × _fraccion`; Σ aportes = `importe_ancestros`."""
    ucs = _cargar_ucs()
    árboles = {
        "cc": cargar_árbol_centros_de_coste(_RUTA_BASE),
        "act": cargar_árbol_actividades(_RUTA_BASE),
        "ec": cargar_árbol_elementos_de_coste(_RUTA_BASE),
    }
    eje = p.eje
    if eje not in _EJES_COLS:
        raise HTTPException(status_code=400, detail=f"eje inválido: {eje!r}")
    col = _EJES_COLS[eje]
    ctx = {
        "cc": p.centros_de_coste, "act": p.actividades, "ec": p.elementos_de_coste,
    }

    # Filtrar por los OTROS ejes (subárbol de su contexto).
    sub = ucs
    for e2 in ("cc", "act", "ec"):
        if e2 == eje:
            continue
        sel = _expandir(ctx[e2], árboles[e2])
        if sel:
            sub = sub.filter(pl.col(_EJES_COLS[e2]).is_in(list(sel)))

    # El roll-down de ancestros se calcula sobre el árbol COMPLETO, anclado en
    # la raíz real (UJI): el coste de cada ancestro se reparte entre TODOS sus
    # descendientes, así que la fracción que llega al nodo depende del árbol
    # entero, no del ámbito de foco. `scope_slugs` se ignora a estos efectos.
    raíces_ámbito = {árboles[eje].raíz.identificador}
    if sub.is_empty():
        return {"n_total": 0, "n_devueltas": 0, "filas": []}

    # Árbol del eje (coste directo por slug) + subárbol directo.
    agg = sub.group_by(col).agg(pl.col("importe").sum().alias("_imp"))
    directo_imp: dict[str, float] = {
        r[col]: float(r["_imp"] or 0.0) for r in agg.iter_rows(named=True) if r[col] is not None
    }
    hijos_de: dict[str, set[str]] = {}
    for s in directo_imp:
        cam = _camino_árbol(árboles[eje], s, raíces_ámbito)
        for padre, hijo in zip(cam, cam[1:]):
            hijos_de.setdefault(padre, set()).add(hijo)

    subtree_d: dict[str, float] = {}

    def _subtree(slug: str) -> float:
        s = directo_imp.get(slug, 0.0)
        for c in hijos_de.get(slug, set()):
            s += _subtree(c)
        subtree_d[slug] = s
        return s

    cam_n = _camino_árbol(árboles[eje], p.slug, raíces_ámbito)
    if cam_n:
        _subtree(cam_n[0])

    # Ratios de cada arista del camino raíz→N: subtree_d[hijo] / Σ subtree_d(hermanos).
    fracs: dict[str, float] = {}  # ancestro estricto Ai -> ρ(Ai, N)
    acc = 1.0
    for j in range(len(cam_n) - 2, -1, -1):  # de N-1 hacia la raíz
        padre, hijo = cam_n[j], cam_n[j + 1]
        hermanos = hijos_de.get(padre, set())
        total = sum(subtree_d.get(h, 0.0) for h in hermanos)
        ratio = (subtree_d.get(hijo, 0.0) / total) if total > 0 else 0.0
        acc *= ratio
        fracs[padre] = acc  # ρ(padre, N)

    filas: list[dict] = []
    for ai, rho in fracs.items():
        if rho <= 0:
            continue
        ucs_ai = sub.filter(pl.col(col) == ai)
        for fila in ucs_ai.iter_rows(named=True):
            fila = dict(fila)
            fila["_fraccion"] = rho
            filas.append(fila)
    # Ordenar por aporte efectivo (importe × fracción) descendente.
    filas.sort(key=lambda f: abs(float(f.get("importe") or 0.0) * f["_fraccion"]), reverse=True)
    n_total = len(filas)
    return {
        "n_total": n_total,
        "n_devueltas": min(n_total, p.limit),
        "filas": filas[: p.limit],
    }


# ----------------------------------------------------------------------
# Configuraciones guardadas: CRUD.
# ----------------------------------------------------------------------

_NOMBRE_OK = re.compile(r"^[\wáéíóúñÁÉÍÓÚÑ.\-\s]{1,80}$")


def _path_config(nombre: str) -> Path:
    if not _NOMBRE_OK.match(nombre):
        raise HTTPException(
            status_code=400,
            detail="Nombre inválido: usa letras, dígitos, espacios, '.' o '-'.",
        )
    return _DIR_CONFIGS / f"{nombre}.yaml"


class ConfigMeta(BaseModel):
    nombre: str


@router.get("/configs", response_model=list[ConfigMeta])
def listar_configs() -> list[ConfigMeta]:
    if not _DIR_CONFIGS.exists():
        return []
    out = []
    for p in sorted(_DIR_CONFIGS.glob("*.yaml")):
        out.append(ConfigMeta(nombre=p.stem))
    return out


@router.get("/configs/{nombre}", response_model=Filtro)
def cargar_config(nombre: str) -> Filtro:
    p = _path_config(nombre)
    if not p.exists():
        raise HTTPException(status_code=404, detail=f"Config {nombre!r} no existe")
    with p.open(encoding="utf-8") as f:
        datos = yaml.safe_load(f) or {}
    return Filtro(**datos)


@router.put("/configs/{nombre}", response_model=ConfigMeta)
def guardar_config(nombre: str, filtro: Filtro) -> ConfigMeta:
    p = _path_config(nombre)
    p.parent.mkdir(parents=True, exist_ok=True)
    with p.open("w", encoding="utf-8") as f:
        yaml.safe_dump(filtro.model_dump(), f, allow_unicode=True, sort_keys=False)
    return ConfigMeta(nombre=nombre)


@router.delete("/configs/{nombre}")
def borrar_config(nombre: str) -> dict[str, str]:
    p = _path_config(nombre)
    if not p.exists():
        raise HTTPException(status_code=404, detail=f"Config {nombre!r} no existe")
    p.unlink()
    return {"borrada": nombre}


# ----------------------------------------------------------------------
# Exportaciones: XLSX y PDF
# ----------------------------------------------------------------------


def _aplanar(raíces: list[NodoJerarquía], depth: int = 0) -> list[tuple[int, NodoJerarquía]]:
    out: list[tuple[int, NodoJerarquía]] = []
    for n in raíces:
        out.append((depth, n))
        out.extend(_aplanar(n.hijos, depth + 1))
    return out


@router.post("/excel")
def exportar_excel(filtro: Filtro) -> StreamingResponse:
    """Devuelve un .xlsx con la estructura jerárquica del informe."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.styles.borders import Border, Side

    res = consulta(filtro)
    wb = Workbook()
    ws = wb.active
    ws.title = "Informe a la carta"
    calibri = Font(name="Calibri", size=11)
    calibri_bold = Font(name="Calibri", size=11, bold=True)
    calibri_title = Font(name="Calibri", size=13, bold=True)
    fill_l1 = PatternFill("solid", fgColor="E7EEF6")
    fill_l2 = PatternFill("solid", fgColor="F2F6FB")
    thin = Side(style="thin", color="888888")
    medio = Side(style="medium", color="333333")

    ws["A1"] = "Informe a la carta"
    ws["A1"].font = calibri_title
    ws.merge_cells("A1:E1")
    ws["A2"] = f"Estructura: {_RAÍZ_LABEL.get(filtro.estructura, filtro.estructura)}"
    ws["A2"].font = Font(name="Calibri", size=10, italic=True)
    ws.merge_cells("A2:E2")

    headers = ["Concepto", "Directo (€)", "Descendientes (€)", "Ancestros (€)", "Total (€)"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=4, column=col, value=h)
        c.font = calibri_bold
        c.border = Border(top=medio, bottom=medio)
        c.alignment = Alignment(horizontal="left" if col == 1 else "right")

    row = 5
    for depth, n in _aplanar(res.raíces):
        sangría = "    " * depth
        etq = f"{sangría}[{n.eje}] {n.código}  {n.descripción}"
        font = calibri_bold if depth == 0 else calibri
        descendientes = round(n.importe - n.importe_directo, 2)
        total = round(n.importe + n.importe_ancestros, 2)
        ws.cell(row=row, column=1, value=etq).font = font
        ws.cell(row=row, column=2, value=n.importe_directo or None).font = font
        ws.cell(row=row, column=3, value=descendientes or None).font = font
        ws.cell(row=row, column=4, value=n.importe_ancestros or None).font = font
        ws.cell(row=row, column=5, value=total).font = font
        for col in (2, 3, 4, 5):
            ws.cell(row=row, column=col).number_format = "#,##0.00"
        if depth == 0:
            for col in range(1, 6):
                ws.cell(row=row, column=col).fill = fill_l1
        elif depth == 1:
            for col in range(1, 6):
                ws.cell(row=row, column=col).fill = fill_l2
        for col in range(1, 6):
            ws.cell(row=row, column=col).border = Border(bottom=thin)
        row += 1

    # Total general (los ancestros se compensan dentro del árbol: el total
    # global del sistema sigue siendo `res.importe`).
    ws.cell(row=row, column=1, value=f"Total ({res.n_ucs:,} UC)".replace(",", ".")).font = calibri_bold
    ws.cell(row=row, column=5, value=res.importe).font = calibri_bold
    ws.cell(row=row, column=5).number_format = "#,##0.00"
    for col in range(1, 6):
        ws.cell(row=row, column=col).border = Border(top=medio, bottom=medio)

    ws.column_dimensions["A"].width = 80
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": 'attachment; filename="informe_a_la_carta.xlsx"',
        },
    )


def _fmt_euro(v: float) -> str:
    """Formato europeo «1.234.567,89» (sin símbolo)."""
    neg = v < 0
    av = abs(v)
    entero = int(av)
    cent = int(round((av - entero) * 100))
    if cent == 100:
        cent = 0
        entero += 1
    s = f"{entero:,}".replace(",", ".")
    signo = "−" if neg else ""
    return f"{signo}{s},{cent:02d}"


@router.post("/pdf")
def exportar_pdf(filtro: Filtro) -> StreamingResponse:
    """Compila un PDF Typst con la estructura jerárquica del informe."""
    res = consulta(filtro)

    filas_typ: list[str] = []
    primer_nivel_1 = True
    for depth, n in _aplanar(res.raíces):
        sangría_pt = depth * 1.2
        etiqueta = (
            f'[#h({sangría_pt}em)#text(fill: gray, size: 8pt)[{n.eje}]'
            f'#h(0.4em)#text(fill: gray)[{_typ_escape(n.código)}]'
            f'#h(0.6em){_typ_escape(n.descripción)}]'
        )
        descendientes = round(n.importe - n.importe_directo, 2)
        total = round(n.importe + n.importe_ancestros, 2)
        dir_celda = f'[{_fmt_euro(n.importe_directo) if n.importe_directo else "—"}]'
        des_celda = f'[{_fmt_euro(descendientes) if descendientes else "—"}]'
        anc_celda = f'[{_fmt_euro(n.importe_ancestros) if n.importe_ancestros else "—"}]'
        tot_celda = f'[{_fmt_euro(total)}]'
        if depth == 0:
            sep = ""
            if not primer_nivel_1:
                sep = "table.hline(stroke: 0.6pt + luma(35%)),\n        "
            primer_nivel_1 = False
            filas_typ.append(
                sep + f"strong({etiqueta}), strong({dir_celda}), strong({des_celda}), strong({anc_celda}), strong({tot_celda}),"
            )
        else:
            filas_typ.append(f"{etiqueta}, {dir_celda}, {des_celda}, {anc_celda}, {tot_celda},")

    cuerpo_tabla = "\n        ".join(filas_typ)
    orden_lbl = _RAÍZ_LABEL.get(filtro.estructura, filtro.estructura)

    typ = f"""\
#set page(paper: "a4", flipped: true, margin: 1.5cm)
#set text(font: "Calibri", lang: "es", size: 9pt)
#set table(
    inset: (x: 0.5em, y: 0.25em),
    fill: (_, row) => if row > 0 and calc.odd(row) {{ luma(95%) }} else {{ none }},
)

#align(center)[
    #text(size: 16pt, weight: "bold")[Informe a la carta]
    #v(0.2cm)
    #text(size: 10pt, fill: gray)[Estructura: {orden_lbl}]
]
#v(0.4cm)

#table(
    columns: (1fr, auto, auto, auto, auto),
    stroke: none,
    align: (left, right, right, right, right),
    table.hline(),
    [*Concepto*], [*Directo (€)*], [*Descendientes (€)*], [*Ancestros (€)*], [*Total (€)*],
    table.hline(),
        {cuerpo_tabla}
    table.hline(),
    strong[Total ({_fmt_int(res.n_ucs)} UC)], [], [], [], strong[{_fmt_euro(res.importe)}],
    table.hline(),
)
"""

    with tempfile.TemporaryDirectory() as tmp:
        tmp_dir = Path(tmp)
        src = tmp_dir / "informe_carta.typ"
        out = tmp_dir / "informe_carta.pdf"
        src.write_text(typ, encoding="utf-8")
        try:
            subprocess.run(
                ["typst", "compile", "--root", str(tmp_dir), str(src), str(out)],
                check=True, timeout=60, capture_output=True,
            )
        except FileNotFoundError:
            raise HTTPException(status_code=500, detail="typst no está disponible en el PATH")
        except subprocess.CalledProcessError as e:
            raise HTTPException(
                status_code=500,
                detail=f"Error compilando PDF: {e.stderr.decode(errors='ignore')[:500]}",
            )
        pdf_bytes = out.read_bytes()

    buf = io.BytesIO(pdf_bytes)
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": 'attachment; filename="informe_a_la_carta.pdf"'},
    )


def _fmt_int(n: int) -> str:
    return f"{int(n):,}".replace(",", ".")


def _typ_escape(s: str) -> str:
    """Escapa caracteres especiales para contenido literal Typst."""
    return s.replace("\\", "\\\\").replace("#", "\\#").replace("[", "\\[").replace("]", "\\]").replace("*", "\\*").replace("_", "\\_").replace("@", "\\@")
